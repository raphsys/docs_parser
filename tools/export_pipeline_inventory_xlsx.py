#!/usr/bin/env python3
"""Export a visual, page-by-page inventory of the PAGEPRINT -> PAGETRANSLATE pipeline.

Produces:
  - an XLSX workbook: one sheet per page + an "Overview" sheet. Each row is an
    extracted element with its granularity (phrase/expression/word/abbreviation
    or formula/table/figure/region), its translatable state, the translation,
    the status and the QA reasons.
  - an assets/ folder per page: the rendered source page, the bbox overlay,
    and the background layer when available.

Input: a directory produced by tools/run_pageprint_pagetranslate_audit.py
(containing input_data_*.json, pagetranslate_result_*.json, source_pages/).
"""

from __future__ import annotations

import argparse
import glob
import json
import re
import shutil
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from tools.audit_translation_selection import _structural, audit_page, classify_granularity
from pagetranslate.quality import unit_quality

HEADER = ["#", "catégorie", "granularité", "role", "object_type", "traduisible",
          "source", "traduction", "statut", "review", "qa_reasons", "bbox"]

FILL_HEADER = PatternFill("solid", fgColor="2F5496")
FILL_YES = PatternFill("solid", fgColor="C6EFCE")
FILL_NO = PatternFill("solid", fgColor="FFC7CE")
FILL_STRUCT = PatternFill("solid", fgColor="FFF2CC")
FONT_HEADER = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
COL_WIDTHS = [5, 18, 13, 18, 16, 12, 60, 60, 12, 8, 26, 26]


def _norm(t: str) -> str:
    return re.sub(r"\s+", " ", str(t or "")).strip().lower()


def _translation_map(pt_result: dict) -> dict:
    out = {}
    for u in pt_result.get("units") or []:
        out[_norm(u.get("source_text"))] = u
    return out


def _profile():
    return {"source_lang": "en", "target_lang": "fr"}


def build_page_rows(input_data: dict, pt_result: dict) -> tuple[list[list], dict]:
    audit = audit_page(input_data)
    st = audit["structural"]
    tmap = _translation_map(pt_result or {})
    rows: list[list] = []
    n = 0

    # 1. Translatable text elements (with translation).
    for it in audit["translatable"]:
        n += 1
        src = it["text"]
        u = tmap.get(_norm(src)) or {}
        translation = u.get("translated_text") or ""
        status = u.get("status") or ""
        review = u.get("needs_review")
        q = unit_quality(src, translation, {}, _profile()) if translation else {}
        rows.append([
            n, "texte", it["granularity"], it.get("role"), it.get("object_type"), "oui",
            src, translation, status, "oui" if review else "", ", ".join(q.get("qa_reasons") or []), "",
        ])

    # 2. Non-translatable text elements.
    for it in audit["non_translatable"]:
        n += 1
        rows.append([
            n, "texte (exclu)", it["granularity"], it.get("role"), it.get("object_type"), "non",
            it["text"], "", "", "", it.get("reason"), "",
        ])

    # 3. Structural / non-textual elements.
    for f in st["formulas"]:
        n += 1
        rows.append([n, "formule", "—", "formula", "formula_expression", "non",
                     f.get("text"), "", f.get("preservation_mode") or "", "", "formula_zone", str(f.get("bbox"))])
    for c in st["code_blocks"]:
        n += 1
        rows.append([n, "code", "—", "code_block", "code", "non", c.get("text"), "", "", "", "code_zone", str(c.get("bbox"))])
    for t in st["tables"]:
        n += 1
        rows.append([n, "table", "—", "table", f"{t.get('cells')} cellules", "non",
                     f"table {t.get('table_id')} — colonnes: {t.get('columns')}", "", "", "", "table_zone", str(t.get("bbox"))])
    for fig in st["figures"]:
        n += 1
        rows.append([n, "figure/image", "—", "figure", f"{fig.get('diagram_labels')} labels", "non",
                     "(zone figure/diagramme)", "", "", "", "figure_zone", str(fig.get("bbox"))])
    for rtype, count in sorted(st["regions"].items(), key=lambda x: -x[1]):
        n += 1
        rows.append([n, "région", "—", rtype, f"x{count}", "non", f"(zone détectée: {rtype})", "", "", "", "special_zone", ""])

    summary = {**audit["summary"], "page_role": audit["page_role"],
               "formulas": len(st["formulas"]), "tables": len(st["tables"]),
               "figures": len(st["figures"]), "background": st["background"]["has_background_layer"],
               "statuses": (pt_result or {}).get("statuses") or {}}
    return rows, summary


def _style_sheet(ws, rows: list[list], title_lines: list[str]) -> None:
    r = 1
    for line in title_lines:
        ws.cell(row=r, column=1, value=line).font = Font(bold=True)
        r += 1
    header_row = r + 1
    for c, name in enumerate(HEADER, start=1):
        cell = ws.cell(row=header_row, column=c, value=name)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
    for i, row in enumerate(rows):
        rr = header_row + 1 + i
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=rr, column=c, value=val)
            cell.alignment = WRAP
        trad = str(row[5])
        cat = str(row[1])
        fill = FILL_YES if trad == "oui" else (FILL_STRUCT if cat not in {"texte", "texte (exclu)"} else FILL_NO)
        ws.cell(row=rr, column=6).fill = fill
    for c, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if rows:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(HEADER))}{header_row + len(rows)}"


def _copy_assets(input_data: dict, audit_dir: Path, tag: str, dest: Path) -> dict:
    dest.mkdir(parents=True, exist_ok=True)
    copied = {}
    assets = input_data.get("assets") or {}
    src_img = assets.get("source_image_path")
    if src_img and Path(src_img).is_file():
        shutil.copy2(src_img, dest / f"page_{Path(src_img).name}")
        copied["page"] = (dest / f"page_{Path(src_img).name}").name
    bbox = audit_dir / f"pageprint_bboxes_{tag}.png"
    if bbox.is_file():
        shutil.copy2(bbox, dest / f"bboxes_{tag}.png")
        copied["bboxes"] = f"bboxes_{tag}.png"
    bg = assets.get("background_path")
    if bg and Path(bg).is_file():
        shutil.copy2(bg, dest / f"background_{Path(bg).name}")
        copied["background"] = (dest / f"background_{Path(bg).name}").name
    else:
        copied["background"] = "non disponible (pas de couche de fond exportée)"
    return copied


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--audit-dir", required=True, help="Directory from run_pageprint_pagetranslate_audit.py")
    parser.add_argument("--out", default=None)
    args = parser.parse_args()

    audit_dir = Path(args.audit_dir)
    out = Path(args.out) if args.out else audit_dir / "pipeline_inventory"
    out.mkdir(parents=True, exist_ok=True)
    assets_root = out / "assets"

    wb = Workbook()
    overview = wb.active
    overview.title = "Overview"
    ov_header = ["page", "page_role", "éléments", "traduisibles", "phrases", "expressions", "mots",
                 "abréviations", "formules", "tables", "figures", "fond", "runtime", "qualité_ling", "santé"]
    for c, name in enumerate(ov_header, start=1):
        cell = overview.cell(row=1, column=c, value=name)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER

    files = sorted(glob.glob(str(audit_dir / "input_data_*.json")))
    ov_r = 2
    used_names = set()
    for path in files:
        tag = Path(path).stem.replace("input_data_", "")
        input_data = json.loads(Path(path).read_text(encoding="utf-8"))
        pt_path = audit_dir / f"pagetranslate_result_{tag}.json"
        pt_result = json.loads(pt_path.read_text(encoding="utf-8")) if pt_path.is_file() else {}

        rows, summary = build_page_rows(input_data, pt_result)
        assets = _copy_assets(input_data, audit_dir, tag, assets_root / tag)

        sheet_name = re.sub(r"[\[\]:*?/\\]", "_", tag)[:28]
        base = sheet_name
        i = 1
        while sheet_name in used_names:
            sheet_name = f"{base[:26]}_{i}"
            i += 1
        used_names.add(sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        gran = summary["translatable_by_granularity"]
        title_lines = [
            f"Page: {tag}    page_role: {summary['page_role']}    santé sélection: {summary['selection_health']}",
            f"Assets: {assets.get('page','-')} | {assets.get('bboxes','-')} | fond: {assets.get('background')}",
            f"Statuts pagetranslate: {summary.get('statuses')}",
        ]
        _style_sheet(ws, rows, title_lines)

        overview.cell(row=ov_r, column=1, value=tag)
        overview.cell(row=ov_r, column=2, value=summary["page_role"])
        overview.cell(row=ov_r, column=3, value=len(rows))
        overview.cell(row=ov_r, column=4, value=summary["translatable_count"])
        overview.cell(row=ov_r, column=5, value=gran.get("phrase", 0))
        overview.cell(row=ov_r, column=6, value=gran.get("expression", 0))
        overview.cell(row=ov_r, column=7, value=gran.get("word", 0))
        overview.cell(row=ov_r, column=8, value=gran.get("abbreviation", 0))
        overview.cell(row=ov_r, column=9, value=summary["formulas"])
        overview.cell(row=ov_r, column=10, value=summary["tables"])
        overview.cell(row=ov_r, column=11, value=summary["figures"])
        overview.cell(row=ov_r, column=12, value="oui" if summary["background"] else "non")
        overview.cell(row=ov_r, column=13, value=summary["statuses"].get("translation_runtime_status"))
        overview.cell(row=ov_r, column=14, value=summary["statuses"].get("linguistic_quality_status"))
        overview.cell(row=ov_r, column=15, value=summary["selection_health"])
        ov_r += 1

    for c in range(1, len(ov_header) + 1):
        overview.column_dimensions[get_column_letter(c)].width = 16
    overview.freeze_panes = "A2"

    xlsx_path = out / "pipeline_inventory.xlsx"
    wb.save(xlsx_path)
    print(f"XLSX : {xlsx_path}")
    print(f"Assets par page : {assets_root}")
    print(f"Pages : {len(files)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
